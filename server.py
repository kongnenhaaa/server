import http.server
import json
import os
import sys
import socket
import subprocess
from datetime import datetime
import requests
import hashlib
import time
import threading
import queue
from socketserver import ThreadingMixIn
import base64

PORT = 5000
TOKEN_PREFIX = "token"

excel_lock = threading.Lock()
print_lock = threading.Lock()

class Colors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def append_token_to_excel(exchanged_token, phone, created_at, excel_file_path):
    with excel_lock:
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            headers = ["Exchanged Token", "Created At", "Phone"]
            
            # Load or create workbook
            if os.path.exists(excel_file_path):
                try:
                    wb = openpyxl.load_workbook(excel_file_path)
                    ws = wb.active
                    
                    # Check what is in the first row
                    first_row_vals = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
                    first_row_cleaned = [v for v in first_row_vals if v is not None]
                    
                    if not first_row_cleaned:
                        # The first row is completely empty. Let's write headers to the first row!
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_idx, value=header)
                    elif first_row_vals[:len(headers)] == headers:
                        # Headers are correct, do nothing
                        pass
                    elif first_row_vals[:3] == ["Exchanged Token", "Phone", "Created At"]:
                        # Swap column 2 and 3 to fix the previous format
                        for row_idx in range(1, ws.max_row + 1):
                            phone_val = ws.cell(row=row_idx, column=2).value
                            created_at_val = ws.cell(row=row_idx, column=3).value
                            ws.cell(row=row_idx, column=2, value=created_at_val)
                            ws.cell(row=row_idx, column=3, value=phone_val)
                    elif first_row_vals[:2] == ["Exchanged Token", "Created At"]:
                        # Upgrade 2-column format to 3-column format
                        ws.cell(row=1, column=3, value="Phone")
                    elif len(first_row_cleaned) >= 4 and "App Name" in first_row_cleaned:
                        # Old format (5 columns). Re-create clean with 3 columns.
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Tokens"
                        ws.append(headers)
                    else:
                        # Header was deleted, but sheet contains data in row 1. Insert header row.
                        ws.insert_rows(1)
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=1, column=col_idx, value=header)
                except Exception:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Tokens"
                    ws.append(headers)
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Tokens"
                ws.append(headers)
    
            # Append new row
            ws.append([exchanged_token, created_at, phone])
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)
                
            try:
                wb.save(excel_file_path)
                with print_lock:
                    print(f" [EXCEL_SUCCESS] Successfully appended token to Excel: {excel_file_path}")
            except PermissionError:
                timestamp_suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
                fallback_path = excel_file_path.replace(".xlsx", f"_backup_{timestamp_suffix}.xlsx")
                wb.save(fallback_path)
                with print_lock:
                    print(f" [EXCEL_WARNING] File tokens.xlsx is open in another app. Saved backup to: {fallback_path}")
        except Exception as e:
            with print_lock:
                print(f" [EXCEL_ERROR] Error saving to Excel: {e}")

def append_json_to_excel(token_data, excel_file_path):
    with excel_lock:
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            if os.path.exists(excel_file_path):
                try:
                    wb = openpyxl.load_workbook(excel_file_path)
                    ws = wb.active
                except Exception:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Tokens JSON"
                    ws.append(["Token JSON Data"])
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Tokens JSON"
                ws.append(["Token JSON Data"])
    
            json_str = json.dumps(token_data, ensure_ascii=False, indent=4)
            
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=1, value=json_str)
            
            ws.column_dimensions['A'].width = 100
            ws.cell(row=row_idx, column=1).alignment = Alignment(wrapText=True)
                
            try:
                wb.save(excel_file_path)
                with print_lock:
                    print(f" [EXCEL_SUCCESS] Successfully appended JSON token to Excel: {os.path.basename(excel_file_path)}")
            except PermissionError:
                timestamp_suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
                fallback_path = excel_file_path.replace(".xlsx", f"_backup_{timestamp_suffix}.xlsx")
                wb.save(fallback_path)
                with print_lock:
                    print(f" [EXCEL_WARNING] File {os.path.basename(excel_file_path)} is open in another app. Saved backup to: {os.path.basename(fallback_path)}")
        except Exception as e:
            with print_lock:
                print(f" [EXCEL_ERROR] Error saving JSON to Excel: {e}")

def exchange_7up_token(zalo_token):
    auth_url = "https://7up-april-utc-public-api.adtimabox.vn/digital-api/digital-user-auth/getTokenDigitalMiniApp"
    campaign_id = "69ddf7b66f593f26c5daf9a5"
    secret_key = "mLzu2c89ZRjkP2bN"
    
    headers_auth = {
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0 (Linux; Android 9; Mi A1 Build/PKQ1.180917.001;) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/138.0.7204.179 Mobile Safari/537.36 Zalo android/260501901 ZaloTheme/light ZaloLanguage/vi",
        "Origin": "https://h5.zdn.vn",
        "Referer": "https://h5.zdn.vn/"
    }
    
    max_retries = 3
    for attempt in range(max_retries):
        timestamp_ms = int(time.time() * 1000)
        sign_str = f"{campaign_id}{secret_key}{timestamp_ms}"
        mac = hashlib.sha256(sign_str.encode('utf-8')).hexdigest()
        
        payload = {
            "campaignId": campaign_id,
            "tokenUser": zalo_token,
            "time": timestamp_ms,
            "mac": mac,
            "channelProvider": "zalo"
        }
        
        try:
            auth_res = requests.post(auth_url, json=payload, headers=headers_auth, timeout=20)
            if auth_res.status_code == 200:
                res_json = auth_res.json()
                data_obj = res_json.get("data")
                if isinstance(data_obj, dict) and data_obj.get("accessToken"):
                    return data_obj.get("accessToken")
                else:
                    with print_lock:
                        print(f" {Colors.WARNING}[EXCHANGE WARN] Lần {attempt+1}: Thiếu accessToken. Res: {res_json}{Colors.ENDC}")
            else:
                with print_lock:
                    print(f" {Colors.WARNING}[EXCHANGE WARN] Lần {attempt+1}: Status {auth_res.status_code}, Body: {auth_res.text}{Colors.ENDC}")
        except Exception as e:
            with print_lock:
                print(f" {Colors.FAIL}[EXCHANGE ERROR] Lần {attempt+1}: {e}{Colors.ENDC}")
                
        if attempt < max_retries - 1:
            time.sleep(2)
            
    return None

def process_token_data(token_data):
    try:
        # Print beautiful headers to the console
        with print_lock:
            print(f"\n{Colors.OKBLUE}{Colors.BOLD}*=========================================================================*{Colors.ENDC}")
            print(f"{Colors.OKBLUE}{Colors.BOLD}|                   [+] PROCESSING QUEUED TOKEN                           |{Colors.ENDC}")
            print(f"{Colors.OKBLUE}{Colors.BOLD}*=========================================================================*{Colors.ENDC}")
        
        # Extract attributes safely
        app_id = token_data.get('app_id') or 'N/A'
        
        # Chỉ nhận token từ 7UP app_id: 3151270984274302494
        if app_id != "3151270984274302494":
            with print_lock:
                print(f" {Colors.WARNING}[IGNORING] Nhận được token từ app khác ({app_id}). Bỏ qua!{Colors.ENDC}")
            return

        app_name = token_data.get('name') or 'Unknown'
        access_token = token_data.get('access_token', '')
        captured_at = token_data.get('captured_at', 0)
        
        time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if captured_at:
            try:
                time_str = datetime.fromtimestamp(float(captured_at) / 1000.0).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                pass
            
        with print_lock:
            print(f" {Colors.BOLD}App Name:      {Colors.ENDC}{Colors.OKGREEN}{app_name}{Colors.ENDC}")
            print(f" {Colors.BOLD}Zalo App ID:   {Colors.ENDC}{Colors.OKCYAN}{app_id}{Colors.ENDC}")
            print(f" {Colors.BOLD}Captured At:   {Colors.ENDC}{time_str}")
            if access_token:
                truncated = access_token[:25] + "..." + access_token[-25:] if len(access_token) > 50 else access_token
                print(f" {Colors.BOLD}Access Token:  {Colors.ENDC}{Colors.WARNING}{truncated}{Colors.ENDC}")
        
        # Add human-readable created_at timestamp
        token_data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Generate a unique, safe filename based on the app name and current timestamp
        safe_app_name = "".join(c if c.isalnum() or c in ('-', '_') else '_' for c in app_name).strip('_')
        if not safe_app_name:
            safe_app_name = "unknown"
        
        # Format timestamp to include YYYYMMDD_HHMMSS and millisecond suffix to ensure uniqueness
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        ms_str = datetime.now().strftime("%f")[:3]
        
        # Resolve path relative to the script's own folder
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Ensure unique filename check
        counter = 1
        token_file_name = f"token_{safe_app_name}_{timestamp_str}_{ms_str}.json"
        token_file_path = os.path.join(script_dir, token_file_name)
        while os.path.exists(token_file_path):
            token_file_name = f"token_{safe_app_name}_{timestamp_str}_{ms_str}_{counter}.json"
            token_file_path = os.path.join(script_dir, token_file_name)
            counter += 1

        # Đổi token sang 7UP JWT
        exchanged_token = None
        if access_token:
            with print_lock:
                print(f" {Colors.OKCYAN}[EXCHANGING] Đang gọi API 7UP để đổi JWT token...{Colors.ENDC}")
            exchanged_token = exchange_7up_token(access_token)
            if exchanged_token:
                with print_lock:
                    print(f" {Colors.OKGREEN}[EXCHANGE SUCCESS] Đổi thành công!{Colors.ENDC}")
                # Không lưu exchanged_token vào token_data để loại bỏ khỏi file json
            else:
                with print_lock:
                    print(f" {Colors.FAIL}[EXCHANGE FAILED] Không thể đổi token.{Colors.ENDC}")

        # Write the token payload to the new unique file
        with open(token_file_path, "w", encoding="utf-8") as f:
            json.dump(token_data, f, indent=4, ensure_ascii=False)
        
        # Export to Excel file
        if exchanged_token:
            phone = ""
            try:
                parts = exchanged_token.split('.')
                if len(parts) >= 2:
                    payload = parts[1]
                    payload += '=' * (-len(payload) % 4)
                    decoded = base64.urlsafe_b64decode(payload).decode('utf-8')
                    payload_json = json.loads(decoded)
                    phone = payload_json.get('phone', '')
            except Exception:
                pass
                
            excel_file_path = os.path.join(script_dir, "tokens.xlsx")
            append_token_to_excel(
                exchanged_token=exchanged_token,
                phone=phone,
                created_at=token_data["created_at"],
                excel_file_path=excel_file_path
            )
        
        json_excel_file_path = os.path.join(script_dir, "tokens_json.xlsx")
        append_json_to_excel(token_data, json_excel_file_path)
        
        with print_lock:
            print(f"{Colors.OKBLUE}---------------------------------------------------------------------------{Colors.ENDC}")
            print(f" [SUCCESS] Successfully wrote token payload to: {Colors.BOLD}{Colors.OKGREEN}{token_file_path}{Colors.ENDC}")
            print(f"{Colors.OKBLUE}*=========================================================================*{Colors.ENDC}\n")

    except Exception as e:
        with print_lock:
            print(f"\n{Colors.FAIL}{Colors.BOLD}[ERROR] Error processing token:{Colors.ENDC} {e}\n")

token_queue = queue.Queue()

def token_worker():
    while True:
        try:
            token_data = token_queue.get()
            if token_data is None:
                break
            process_token_data(token_data)
            time.sleep(1)  # Đợi 1s sau khi đổi xong mới chạy đổi tiếp
        except Exception as e:
            with print_lock:
                print(f" [WORKER ERROR] Error in token worker: {e}")
        finally:
            token_queue.task_done()

# Start background processing worker thread
threading.Thread(target=token_worker, daemon=True).start()

class TokenListenerHandler(http.server.BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        # Override to suppress default HTTP logs to keep console clean and beautiful
        pass

    def do_POST(self):
        content_length = int(self.headers.get('Content-Length', 0))
        post_data = self.rfile.read(content_length)
        
        try:
            # Parse the incoming token JSON
            token_data = json.loads(post_data.decode('utf-8'))
            
            with print_lock:
                print(f" {Colors.OKCYAN}[QUEUED] Received token. Adding to queue...{Colors.ENDC}")
            
            token_queue.put(token_data)
            
            # Send successful response immediately to avoid connection timeouts
            try:
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "success", "message": "Token received and queued"}).encode('utf-8'))
            except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
                pass
            
        except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
            pass
        except Exception as e:
            with print_lock:
                print(f"\n{Colors.FAIL}{Colors.BOLD}[ERROR] Error enqueueing token:{Colors.ENDC} {e}\n")
            try:
                self.send_response(400)
                self.send_header('Content-Type', 'application/json')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode('utf-8'))
            except (ConnectionAbortedError, ConnectionResetError, BrokenPipeError):
                pass

    def do_OPTIONS(self):
        # Enable CORS for browser-based testing
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

def get_local_ips():
    ips = []
    # Method 1: Connect to an external socket to detect active local interface
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ips.append(s.getsockname()[0])
        s.close()
    except Exception:
        pass
        
    # Method 2: Fallback/Alternative interfaces
    try:
        hostname = socket.gethostname()
        for info in socket.getaddrinfo(hostname, None):
            ip = info[4][0]
            # Match common private networks (192.168.x.x, 10.x.x.x, 172.x.x.x)
            if ip.startswith("192.") or ip.startswith("10.") or ip.startswith("172."):
                if ip not in ips:
                    ips.append(ip)
    except Exception:
        pass
        
    if not ips:
        ips.append("127.0.0.1")
    return ips

def setup_adb_reverse(port):
    """Setup ADB reverse port forwarding so phone can reach PC server via 127.0.0.1 through USB."""
    try:
        result = subprocess.run(
            ["adb", "reverse", f"tcp:{port}", f"tcp:{port}"],
            capture_output=True, text=True, timeout=10
        )
        if result.returncode == 0:
            print(f" {Colors.OKGREEN}[ADB] Reverse proxy OK!{Colors.ENDC} Phone localhost:{port} → PC localhost:{port}")
            print(f" {Colors.BOLD} ADB Webhook URL:{Colors.ENDC} {Colors.BOLD}{Colors.OKGREEN}http://127.0.0.1:{port}/token{Colors.ENDC}")
            return True
        else:
            stderr = result.stderr.strip()
            print(f" {Colors.WARNING}[ADB] Reverse proxy failed: {stderr}{Colors.ENDC}")
            return False
    except FileNotFoundError:
        print(f" {Colors.WARNING}[ADB] adb not found in PATH. Skipping reverse proxy setup.{Colors.ENDC}")
        return False
    except subprocess.TimeoutExpired:
        print(f" {Colors.WARNING}[ADB] adb reverse timed out. Is a device connected?{Colors.ENDC}")
        return False
    except Exception as e:
        print(f" {Colors.WARNING}[ADB] Error: {e}{Colors.ENDC}")
        return False

class ThreadingHTTPServer(ThreadingMixIn, http.server.HTTPServer):
    daemon_threads = True

def init_excel_file():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file_path = os.path.join(script_dir, "tokens.xlsx")
    if not os.path.exists(excel_file_path):
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            ws = wb.active
            ws.title = "Tokens"
            headers = ["Exchanged Token", "Created At", "Phone"]
            ws.append(headers)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)
                
            wb.save(excel_file_path)
            with print_lock:
                print(f" {Colors.OKGREEN}[EXCEL_INIT] Tự động tạo file Excel mới: {excel_file_path}{Colors.ENDC}")
        except ImportError:
            with print_lock:
                print(f" {Colors.WARNING}[EXCEL_WARNING] openpyxl module not found. Vui lòng chạy 'pip install openpyxl'{Colors.ENDC}")
        except Exception as e:
            with print_lock:
                print(f" {Colors.FAIL}[EXCEL_ERROR] Lỗi khi tạo file Excel: {e}{Colors.ENDC}")

def run(port=PORT):
    server_address = ('', port)
    httpd = ThreadingHTTPServer(server_address, TokenListenerHandler)
    
    print(f"\n{Colors.OKGREEN}{Colors.BOLD}*=============================================================*{Colors.ENDC}")
    print(f"{Colors.OKGREEN}{Colors.BOLD}|          *** ZALO TOKEN WEBHOOK LISTENING SERVER            |{Colors.ENDC}")
    print(f"{Colors.OKGREEN}{Colors.BOLD}*=============================================================*{Colors.ENDC}")
    print(f" Running on port:  {Colors.BOLD}{Colors.OKCYAN}{port}{Colors.ENDC}")
    print(f" Saving tokens to: {Colors.BOLD}{Colors.WARNING}token_[AppName]_[DateTime]_[ms].json{Colors.ENDC}")
    
    # Setup ADB reverse proxy (USB tunnel)
    print(f"{Colors.OKGREEN}---------------------------------------------------------------{Colors.ENDC}")
    adb_ok = setup_adb_reverse(port)
    
    # Auto-detect and print local IP address(es)
    print(f"{Colors.OKGREEN}---------------------------------------------------------------{Colors.ENDC}")
    local_ips = get_local_ips()
    for idx, ip in enumerate(local_ips):
        label = " WiFi Webhook URL:" if idx == 0 else " Or use alternative URL:"
        print(f"{label} {Colors.BOLD}{Colors.OKGREEN}http://{ip}:{port}/token{Colors.ENDC}")
    
    if adb_ok:
        print(f"\n {Colors.OKCYAN}{Colors.BOLD}[TIP]{Colors.ENDC} Dùng ADB (USB): set webhook URL trên app = {Colors.BOLD}http://127.0.0.1:{port}/token{Colors.ENDC}")
        
    print(f"{Colors.OKGREEN}---------------------------------------------------------------{Colors.ENDC}")
    
    # Khởi tạo file excel nếu chưa có
    init_excel_file()
    
    print(" Awaiting incoming tokens...\n")
    
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        # Cleanup ADB reverse on exit
        try:
            subprocess.run(["adb", "reverse", "--remove", f"tcp:{port}"], capture_output=True, timeout=5)
        except Exception:
            pass
        print(f"\n {Colors.WARNING}Shutting down server gracefully...{Colors.ENDC}\n")
        sys.exit(0)

if __name__ == '__main__':
    # Force UTF-8 encoding for stdout on Windows to prevent UnicodeEncodeError with box drawing characters
    if sys.stdout.encoding != 'utf-8':
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except Exception:
            pass
    run()
